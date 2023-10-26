import math
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook
from pandas import DataFrame

from utils.args_manager import ArgsManager
from utils.constant import Constant


class CommunicationComparisonGenerator:
    def __init__(self, args: any, compare_result_data: DataFrame):
        self._args = args
        self._args_manager = ArgsManager()
        self._compare_result_data = compare_result_data
        self._row_index = 3

    def create_sheet(self, workbook: Workbook):
        ws = workbook.create_sheet("CommunicationCompare", 0)
        ws.sheet_properties.tabColor = Constant.YELLOW_COLOR
        base_headers = Constant.CMP_COMMUNICATION_HEADER
        comparison_headers = Constant.CMP_COMMUNICATION_HEADER
        headers = base_headers + comparison_headers + [Constant.DIFF]
        self.set_header(ws, base_headers, comparison_headers, headers)
        self.write_lines(ws, base_headers, comparison_headers, headers)

    def set_header(self, ws, base_headers, comparison_headers, headers):
        # write headers
        base_trace_start_column = 0
        comparison_trace_start_column = len(base_headers)
        diff_start_column = len(base_headers) + len(comparison_headers)
        for col_index, header_name in enumerate(headers):
            ws.cell(row=1, column=col_index + 1).border = Constant.BORDER
            ws.cell(row=1, column=col_index + 1).font = Font(name='Arial')
            ws.cell(row=1, column=col_index + 1).fill = Constant.HEADERS_FILL
            ws.cell(row=2, column=col_index + 1).border = Constant.BORDER
            ws.cell(row=2, column=col_index + 1).font = Font(name='Arial', bold=True)
            ws.cell(row=2, column=col_index + 1).fill = Constant.HEADERS_FILL
            if col_index < comparison_trace_start_column:
                ws.cell(row=1, column=col_index + 1).value = Constant.BASE_PROFILING
            elif col_index < diff_start_column:
                ws.cell(row=1, column=col_index + 1).value = Constant.COMPARISON_PROFILING
            else:
                ws.cell(row=1, column=col_index + 1).value = header_name
            ws.cell(row=2, column=col_index + 1).value = header_name
            dim = ws.cell(row=2, column=col_index + 1).coordinate
            ws.column_dimensions[dim[0]].width = Constant.COLUMN_WIDTH_CLL.get(header_name)
        ws.merge_cells(start_row=1, start_column=base_trace_start_column + 1,
                       end_row=1, end_column=comparison_trace_start_column)
        ws.merge_cells(start_row=1, start_column=comparison_trace_start_column + 1,
                       end_row=1, end_column=diff_start_column)
        ws.merge_cells(start_row=1, start_column=headers.index(Constant.DIFF) + 1,
                       end_row=2, end_column=headers.index(Constant.DIFF) + 1)

    def write_lines(self, ws, base_headers, comparison_headers, headers):
        # write lines
        self._row_index = 3
        for _, row in self._compare_result_data.iterrows():
            self.write_summary_lines(ws, row, headers)
            self._row_index += 1
            self.write_detail_lines(ws, row, base_headers, comparison_headers, headers)
    
    def write_summary_lines(self, ws, row, headers):
        # write summary lines
        base_name = Constant.NA if math.isnan(row[Constant.BASE_CALLS]) else row[Constant.OP_KEY]
        comparison_name = Constant.NA if math.isnan(row[Constant.COMPARISON_CALLS]) else row[Constant.OP_KEY]
        if math.isnan(row[Constant.BASE_SUM]) or math.isnan(row[Constant.COMPARISON_SUM]) or row[
            Constant.BASE_SUM] == 0:
            diff = Constant.NA
        else:
            diff = (row[Constant.COMPARISON_SUM] - row[Constant.BASE_SUM]) / row[Constant.BASE_SUM]
        row_data = [base_name, Constant.NA, row[Constant.BASE_CALLS], row[Constant.BASE_SUM],
                    row[Constant.BASE_AVG], row[Constant.BASE_MAX], row[Constant.BASE_MIN], comparison_name,
                    Constant.NA, row[Constant.COMPARISON_CALLS], row[Constant.COMPARISON_SUM],
                    row[Constant.COMPARISON_AVG], row[Constant.COMPARISON_MAX], row[Constant.COMPARISON_MIN], diff]
        for index, header_name in enumerate(headers):
            if header_name in (
                    Constant.CALLS, Constant.TOTAL_DURATION, Constant.AVG_DURATION, Constant.MAX_DURATION,
                    Constant.MIN_DURATION):
                ws.cell(row=self._row_index, column=index + 1).number_format = '0.00'
            if header_name == Constant.DIFF:
                ws.cell(row=self._row_index, column=index + 1).number_format = '0.00%'
                if diff != Constant.NA and diff < 0:
                    ws.cell(row=self._row_index, column=index + 1).font = Font(name='Arial',
                                                                            color=Constant.GREEN_COLOR)
                elif diff != Constant.NA and diff >= 0:
                    ws.cell(row=self._row_index, column=index + 1).font = Font(name='Arial', color=Constant.RED_COLOR)
            else:
                bold = header_name == Constant.COMMUNICAT_OP
                ws.cell(row=self._row_index, column=index + 1).font = Font(name='Arial', bold=bold)
            value = row_data[index]
            if value != Constant.NA:
                ws.cell(row=self._row_index, column=index + 1).value = value
            ws.cell(row=self._row_index, column=index + 1).border = Constant.BORDER
            ws.cell(row=self._row_index, column=index + 1).fill = PatternFill("solid",
                                                                        fgColor=Constant.SUMMARY_LINE_COLOR)

    def write_detail_lines(self, ws, row, base_headers, comparison_headers, headers):
        # write detail lines
        base_name = Constant.NA if math.isnan(row[Constant.BASE_CALLS]) else row[Constant.OP_KEY]
        comparison_name = Constant.NA if math.isnan(row[Constant.COMPARISON_CALLS]) else row[Constant.OP_KEY]
        base_task_list = self._args_manager.base_profiling.communication_task_data.get(base_name, [])
        comparison_task_list = self._args_manager.comparison_profiling.communication_task_data.get(comparison_name, [])
        if base_task_list:
            base_data = [[data.get("name", ""), float(data.get("dur", 0))] for data in base_task_list]
            base_df = pd.DataFrame(base_data, columns=[Constant.OP_KEY, Constant.DEVICE_DUR])
            base_data = base_df.groupby(Constant.OP_KEY).agg(
                ["count", "sum", "mean", "max", "min"]).reset_index().values.tolist()
        else:
            base_data = []
        if comparison_task_list:
            comparison_data = [[data.get("name", ""), float(data.get("dur", 0))] for data in comparison_task_list]
            comparison_df = pd.DataFrame(comparison_data, columns=[Constant.OP_KEY, Constant.DEVICE_DUR])
            comparison_data = comparison_df.groupby(Constant.OP_KEY).agg(
                ["count", "sum", "mean", "max", "min"]).reset_index().values.tolist()
        else:
            comparison_data = []
        for index in range(max(len(base_data), len(comparison_data))):
            base_detail_data, comparison_detail_data = [Constant.NA] * len(base_headers), [Constant.NA] * len(comparison_headers)
            base_detail_data[0] = "|"
            comparison_detail_data[0] = "|"
            if index < len(base_data):
                total_dur = sum([data[2] for data in base_data])
                percent = 0.0 if total_dur < Constant.EPS else base_data[index][2] / total_dur
                dur_percent = "%.2f%%" % (percent * 100)
                base_data[index][0] = f"{base_data[index][0]} ({dur_percent})"
                base_detail_data[1:] = base_data[index]
            if index < len(comparison_data):
                total_dur = sum([data[2] for data in comparison_data])
                percent = 0.0 if total_dur < Constant.EPS else comparison_data[index][2] / total_dur
                dur_percent = "%.2f%%" % (percent * 100)
                comparison_data[index][0] = f"{comparison_data[index][0]} ({dur_percent})"
                comparison_detail_data[1:] = comparison_data[index]

            detail_data = base_detail_data + comparison_detail_data + [Constant.NA]
            for colum_index in range(len(headers)):
                if headers[colum_index] in (
                        Constant.CALLS, Constant.TOTAL_DURATION, Constant.AVG_DURATION, Constant.MAX_DURATION,
                        Constant.MIN_DURATION):
                    ws.cell(row=self._row_index, column=colum_index + 1) .number_format = '0.00'
                value = detail_data[colum_index]
                if value != Constant.NA:
                    ws.cell(row=self._row_index, column=colum_index + 1).value = value
                    bold = headers[colum_index] == Constant.OP_NAME
                    ws.cell(row=self._row_index, column=colum_index + 1).font = Font(name='Arial', bold=bold)
                ws.cell(row=self._row_index, column=colum_index + 1).border = Constant.BORDER
                if headers[colum_index] == Constant.COMMUNICAT_OP:
                    ws.cell(row=self._row_index, column=colum_index + 1).alignment = Alignment(horizontal="center", vertical="center")
            self._row_index += 1
