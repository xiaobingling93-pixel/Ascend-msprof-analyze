import copy
from collections import namedtuple

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.workbook import Workbook

from utils.args_manager import ArgsManager
from utils.constant import Constant
from utils.tree_builder import TreeBuilder


class OpComparisonGenerator:
    def __init__(self, args: any, compare_result_data: list, compare_type: str):
        self._args = args
        self._compare_result_data = compare_result_data
        self._compare_type = compare_type
        self._base_headers = []
        self._comparison_headers = []
        self._row_index = 3
        self.update_headers()

    def update_headers(self):
        base_profiling_type = ArgsManager().base_profiling_type
        comparison_profiling_type = ArgsManager().comparison_profiling_type
        if self._compare_type == Constant.MEMORY_COMPARE:
            self._base_headers = Constant.CMP_MEMORY_HEADER
            self._comparison_headers = Constant.CMP_MEMORY_HEADER
        elif self._compare_type == Constant.OPERATOR_COMPARE:
            self._base_headers = Constant.GPU_CMP_KERNEL_HEADER if base_profiling_type == Constant.GPU else \
                Constant.NPU_CMP_KERNEL_HEADER
            self._comparison_headers = Constant.GPU_CMP_KERNEL_HEADER if comparison_profiling_type == Constant.GPU \
                else Constant.NPU_CMP_KERNEL_HEADER

    def create_sheet(self, workbook: Workbook):
        ws = workbook.create_sheet(self._compare_type, 0)
        ws.sheet_properties.tabColor = Constant.YELLOW_COLOR
        headers = self._base_headers + self._comparison_headers + [Constant.DIFF, Constant.OP_NAME_FILTER,
                                                                   Constant.DIFF_FILTER]
        self.writer_headers(ws, headers)

        # write lines
        self._row_index = 3
        for data in self._compare_result_data:
            base_event_list = TreeBuilder.get_total_compare_event(data[0], self._compare_type) if data[0] else []
            comparison_event_list = TreeBuilder.get_total_compare_event(data[1], self._compare_type) if data[1] else []
            diff = self.write_summary_lines(ws, headers, data, base_event_list, comparison_event_list)
            self._row_index += 1
            EventListWrapper = namedtuple('EventListWrapper', ['base_event_list', 'comparison_event_list'])
            event_list = EventListWrapper(base_event_list, comparison_event_list)
            self.write_detail_lines(ws, headers, data, diff, event_list)

    def writer_headers(self, ws, headers):
        # write headers
        base_trace_start_column = 0
        comparison_trace_start_column = len(self._base_headers)
        diff_start_column = len(self._base_headers) + len(self._comparison_headers)

        for col_index in range(len(headers)):
            ws.cell(row=1, column=col_index + 1).border = Constant.BORDER
            ws.cell(row=1, column=col_index + 1).font = Font(name='Arial')
            ws.cell(row=1, column=col_index + 1).fill = Constant.HEADERS_FILL
            ws.cell(row=2, column=col_index + 1).border = Constant.BORDER
            ws.cell(row=2, column=col_index + 1).font = Font(name='Arial', bold=True)
            ws.cell(row=2, column=col_index + 1).fill = Constant.HEADERS_FILL
            header_name = headers[col_index]
            if col_index < comparison_trace_start_column:
                ws.cell(row=1, column=col_index + 1).value = Constant.BASE_PROFILING
            elif col_index < diff_start_column:
                ws.cell(row=1, column=col_index + 1).value = Constant.COMPARISON_PROFILING
            else:
                ws.cell(row=1, column=col_index + 1).value = header_name
            ws.cell(row=2, column=col_index + 1).value = header_name
            dim = ws.cell(row=2, column=col_index + 1).coordinate
            width = Constant.COLUMN_WIDTH.get(header_name) if Constant.COLUMN_WIDTH.get(
                header_name) else Constant.DEFAULT_WIDTH
            ws.column_dimensions[dim[0]].width = width
        ws.merge_cells(start_row=1, start_column=base_trace_start_column + 1,
                       end_row=1, end_column=comparison_trace_start_column)
        ws.merge_cells(start_row=1, start_column=comparison_trace_start_column + 1,
                       end_row=1, end_column=diff_start_column)
        ws.merge_cells(start_row=1, start_column=headers.index(Constant.DIFF) + 1,
                       end_row=2, end_column=headers.index(Constant.DIFF) + 1)
        ws.merge_cells(start_row=1, start_column=headers.index(Constant.OP_NAME_FILTER) + 1,
                       end_row=2, end_column=headers.index(Constant.OP_NAME_FILTER) + 1)
        ws.merge_cells(start_row=1, start_column=headers.index(Constant.DIFF_FILTER) + 1,
                       end_row=2, end_column=headers.index(Constant.DIFF_FILTER) + 1)

    def write_summary_lines(self, ws, headers, data, base_event_list, comparison_event_list):
        def ws_write_diff(ws, index, value):
            ws.cell(row=self._row_index, column=index + 1).number_format = '0.00%'
            if value != Constant.NA and value < 0:
                ws.cell(row=self._row_index, column=index + 1).font = Font(name='Arial', color=Constant.GREEN_COLOR)
            elif value != Constant.NA and value >= 0:
                ws.cell(row=self._row_index, column=index + 1).font = Font(name='Arial', color=Constant.RED_COLOR)
        
        def ws_write_diff_filter(ws, index, diff_value):
            if diff_value != Constant.NA and diff_value < 0:
                ws.cell(row=self._row_index, column=index + 1).fill = PatternFill("solid",
                                                                            fgColor=Constant.GREEN_COLOR)
            elif diff_value != Constant.NA and diff_value >= 0:
                ws.cell(row=self._row_index, column=index + 1).fill = PatternFill("solid", fgColor=Constant.RED_COLOR)
        # write summary lines
        base_summary_data, comparison_summary_data = [Constant.NA] * len(self._base_headers), \
                                                        [Constant.NA] * len(self._comparison_headers)
        if data[0]:
            base_summary_data[0] = data[0].name
            base_summary_data[1] = data[0].input_shape
            base_summary_data[2] = data[0].input_type
            base_summary_data[3] = sum(
                [x.compare_index for x in base_event_list]) if base_event_list else Constant.NA
        if data[1]:
            comparison_summary_data[0] = data[1].name
            comparison_summary_data[1] = data[1].input_shape
            comparison_summary_data[2] = data[1].input_type
            comparison_summary_data[3] = sum(
                [x.compare_index for x in comparison_event_list]) if comparison_event_list else Constant.NA
        if base_event_list and comparison_event_list and base_summary_data[3]:
            diff = (comparison_summary_data[3] - base_summary_data[3]) / base_summary_data[3]
        else:
            diff = Constant.NA
        op_name = data[0].name if data[0] else data[1].name

        summary_data = base_summary_data + comparison_summary_data + [diff, op_name, Constant.NA]
        if len(summary_data) < len(headers):
            raise RuntimeError("Fail to write summary lines!")
        for index, header_name in enumerate(headers):
            value = summary_data[index]
            if header_name == Constant.DIFF:
                ws_write_diff(ws, index, value)
            if header_name == Constant.DIFF_FILTER:
                diff_value = summary_data[headers.index(Constant.DIFF)]
                ws_write_diff_filter(ws, index, diff_value)
            elif header_name != Constant.OP_NAME_FILTER:
                ws.cell(row=self._row_index, column=index + 1).fill = PatternFill("solid",
                                                                            fgColor=Constant.SUMMARY_LINE_COLOR)

            if value != Constant.NA:
                ws.cell(row=self._row_index, column=index + 1).value = value
                bold = header_name == Constant.OP_NAME
                if header_name != Constant.DIFF:
                    ws.cell(row=self._row_index, column=index + 1).font = Font(name='Arial', bold=bold)
            ws.cell(row=self._row_index, column=index + 1).border = Constant.BORDER
        return diff

    def write_detail_lines(self, ws, headers, data, diff, event_list):
        def ws_write_helper(ws, colum_index, value, diff, headers):
            if value != Constant.NA:
                ws.cell(row=self._row_index, column=colum_index + 1).value = value
                bold = headers[colum_index] == Constant.OP_NAME
                ws.cell(row=self._row_index, column=colum_index + 1).font = Font(name='Arial', bold=bold)
            ws.cell(row=self._row_index, column=colum_index + 1).border = Constant.BORDER
            if headers[colum_index] == Constant.DIFF_FILTER:
                if diff != Constant.NA and diff < 0:
                    ws.cell(row=self._row_index, column=colum_index + 1).fill = PatternFill("solid",
                                                                                        fgColor=Constant.GREEN_COLOR)
                elif diff != Constant.NA and diff >= 0:
                    ws.cell(row=self._row_index, column=colum_index + 1).fill = PatternFill("solid",
                                                                                        fgColor=Constant.RED_COLOR)
            if headers[colum_index] == Constant.OP_NAME:
                ws.cell(row=self._row_index, column=colum_index + 1).alignment = Alignment(horizontal="center",
                                                                                        vertical="center")

        base_event_list = event_list.base_event_list
        comparison_event_list = event_list.comparison_event_list
        # write detail lines
        op_name = data[0].name if data[0] else data[1].name
        base_event_num, comparison_event_num = len(base_event_list), len(comparison_event_list)
        for index in range(max(base_event_num, comparison_event_num)):
            base_detail_data, comparison_detail_data = [Constant.NA] * len(self._base_headers), \
                                                        [Constant.NA] * len(self._comparison_headers)
            base_detail_data[0] = "|"
            comparison_detail_data[0] = "|"
            if index < base_event_num:
                base_event = base_event_list[index]
                base_detail_data[1:] = base_event.get_record()
            if index < comparison_event_num:
                comparison_event = comparison_event_list[index]
                comparison_detail_data[1:] = comparison_event.get_record()

            detail_data = base_detail_data + comparison_detail_data + [Constant.NA, op_name, Constant.NA]
            for colum_index in range(len(headers)):
                value = detail_data[colum_index]
                ws_write_helper(ws, colum_index, value, diff, headers)
            self._row_index += 1
