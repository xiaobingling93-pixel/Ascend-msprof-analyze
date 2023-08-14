import os

from openpyxl.workbook import Workbook

from comparator.index_comparator import IndexComparator
from comparator.op_comparator import OpComparator
from generation.communication_comparison_generator import CommunicationComparisonGenerator
from generation.op_comparison_generator import OpComparisonGenerator
from utils.constant import Constant
from utils.args_manager import ArgsManager


class ComparisonGenerator:
    def __init__(self, args: any):
        self._args = args
        self._args_manager = ArgsManager()

    def create_excel(self, file_path: str):
        wb = Workbook()
        if not self._args.disable_operator_compare or not self._args.disable_memory_compare:
            op_compare_result = OpComparator(self._args).compare()
            if op_compare_result:
                if not self._args.disable_operator_compare:
                    OpComparisonGenerator(self._args, op_compare_result, Constant.OPERATOR_COMPARE).create_sheet(wb)
                if not self._args.disable_memory_compare:
                    OpComparisonGenerator(self._args, op_compare_result, Constant.MEMORY_COMPARE).create_sheet(wb)

        if not self._args.disable_communication_compare:
            index_compare_result = IndexComparator(self._args).compare()
            if not index_compare_result.empty:
                CommunicationComparisonGenerator(self._args, index_compare_result).create_sheet(wb)

        wb.save(file_path)
        wb.close()
        os.chmod(file_path, Constant.FILE_AUTHORITY)
