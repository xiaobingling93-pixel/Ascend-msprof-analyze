import argparse
import ast
import copy
import json
import os.path
import time
from queue import Queue
import numpy as np

from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.workbook import Workbook

GPU = 0
NPU = 1
NA = 'N/A'
LIMIT_KERNEL = 3
OP_NAME = 'Operator Name'
INPUT_SHAPE = 'Input Shape'
INPUT_TYPE = 'Input Type'
KERNEL_NAME = 'Kernel Name'
DEVICE_DUR = 'Device Duration(us)'
TASK_ID = 'Task Id'
KERNEL_TYPE = 'Kernel Type'
DIFF = 'DIFF: (sum(Trace2 Duration)-sum(Trace1 Duration))/sum(Trace1 Duration)'
OP_NAME_FILTER = 'Operator Name Filter'
DIFF_FILTER = 'DIFF Filter'
BASE_TRACE = 'Base Trace'
COMPARISON_TRACE = 'Comparison Trace'
BASE_TRACE_TYPE = None
COMPARISON_TRACE_TYPE = None
BASE_TYPE = 1
COMPARISON_TYPE = 2
GPU_HEADER = [OP_NAME, INPUT_SHAPE, INPUT_TYPE, KERNEL_NAME, DEVICE_DUR]
NPU_HEADER = [OP_NAME, INPUT_SHAPE, INPUT_TYPE, KERNEL_NAME, TASK_ID, KERNEL_TYPE, DEVICE_DUR]
FILL_DICT = {
    BASE_TYPE: PatternFill("solid", fgColor='003366FF'), COMPARISON_TYPE: PatternFill("solid", fgColor='0033CCCC'),
    DIFF: PatternFill("solid", fgColor='00FF0000'), OP_NAME_FILTER: PatternFill("solid", fgColor='00FFFF00'),
    DIFF_FILTER: PatternFill("solid", fgColor='00FFFF00')
}
COLUMN_WIDTH = {OP_NAME: 50, INPUT_SHAPE: 25, INPUT_TYPE: 25, KERNEL_NAME: 25, DEVICE_DUR: 25,
                TASK_ID: 20, KERNEL_TYPE: 25, DIFF: 25, OP_NAME_FILTER: 25, DIFF_FILTER: 25}
BORDER = Border(top=Side(border_style="thin", color='00000000'),
                left=Side(border_style="thin", color='00000000'),
                right=Side(border_style="thin", color='00000000'),
                bottom=Side(border_style="thin", color='00000000'))


class TorchOpNode:
    def __init__(self, event=None, parent_node=None):
        self._event = event
        self._parent_node = parent_node
        self._child_nodes = []
        self._kernel_list = []
        self._kernel_num = 0

    @property
    def start_time(self):
        return self._event.get("ts", 0)

    @property
    def end_time(self):
        return self._event.get("ts", 0) + self._event.get("dur", 0)

    @property
    def name(self):
        return str(self._event.get("name", NA))

    @property
    def input_shape(self):
        return str(self._event.get("args", {}).get("Input Dims", NA))

    @property
    def input_type(self):
        return str(self._event.get("args", {}).get("Input type", NA))

    @property
    def parent(self):
        return self._parent_node

    @property
    def child_nodes(self):
        return self._child_nodes

    @property
    def kernel_list(self):
        return self._kernel_list

    @property
    def kernel_num(self):
        return self._kernel_num

    def add_child_node(self, child_node):
        self._child_nodes.append(child_node)

    def set_kernel_list(self, kernel_list: list):
        self._kernel_list = kernel_list

    def add_kernel_num(self, kernel_num: int):
        self._kernel_num += kernel_num

    def is_step_profiler(self) -> bool:
        return self.name.find("ProfilerStep#") != -1


class TreeBuilder:
    @classmethod
    def build_tree(cls, event_list: list, flow_kernel_dict: dict) -> TorchOpNode:
        root_node = TorchOpNode()
        event_list.sort(key=lambda x: x.get("ts", 0))
        last_node = root_node
        for event in event_list:
            kernel_list = flow_kernel_dict.get(event.get("ts", 0), [])
            while last_node:
                if last_node == root_node or event.get("ts", 0) < last_node.end_time:
                    tree_node = TorchOpNode(event, last_node)
                    last_node.add_child_node(tree_node)
                    if kernel_list:
                        tree_node.set_kernel_list(kernel_list)
                    last_node = tree_node
                    break
                last_node = last_node.parent
        return root_node

    @classmethod
    def mark_kernel_num(cls, root_node: TorchOpNode, flow_kernel_dict: dict):
        for ts, kernel_list in flow_kernel_dict.items():
            curr_node = root_node
            while curr_node.child_nodes:
                for node in curr_node.child_nodes:
                    if node.start_time <= ts <= node.end_time:
                        node.add_kernel_num(len(kernel_list))
                        curr_node = node
                        break

    @classmethod
    def get_total_kernels(cls, root_node: TorchOpNode) -> list:
        result_list = []
        node_queue = Queue()
        for child_node in root_node.child_nodes:
            node_queue.put(child_node)
        while not node_queue.empty():
            tree_node = node_queue.get()
            result_list.extend(tree_node.kernel_list)
            for child_node in tree_node.child_nodes:
                node_queue.put(child_node)
        return result_list


def read_json_file(file_path: str, trace_type: int) -> any:
    event_list = []
    flow_kernel_dict = {}
    if not os.path.isfile(file_path):
        raise RuntimeError(f"File not exists: {file_path}")
    try:
        with open(file_path, "rt") as file:
            json_data = json.loads(file.read())
    except Exception:
        raise RuntimeError(f"Can't read file: {file_path}")
    flow_start_dict, flow_end_dict, event_dict = {}, {}, {}
    flow_cat = ("async_gpu", "ac2g", "async_npu")
    if trace_type == BASE_TYPE:
        global BASE_TRACE_TYPE
        BASE_TRACE_TYPE = GPU if isinstance(json_data, dict) else NPU
        _type = BASE_TRACE_TYPE
    else:
        global COMPARISON_TRACE_TYPE
        COMPARISON_TRACE_TYPE = GPU if isinstance(json_data, dict) else NPU
        _type = COMPARISON_TRACE_TYPE
    total_events = json_data.get("traceEvents", []) if _type == GPU else json_data
    for event in total_events:
        if event.get("cat") == "cpu_op" or event.get("cat") in ("Runtime", "cuda_runtime"):
            event_list.append(event)
        elif event.get("cat") in flow_cat and event.get("ph") == "s":
            flow_start_dict[event.get("id")] = event
        elif event.get("cat") in flow_cat and event.get("ph") == "f":
            flow_end_dict[event.get("id")] = event
        elif _type == GPU and event.get("cat", "").capitalize() == "Kernel".capitalize():
            event_dict["{}-{}-{}".format(event.get("pid"), event.get("tid"), event.get("ts"))] = event
        elif _type == NPU and event.get("ph") != "f":
            event_dict["{}-{}-{}".format(event.get("pid"), event.get("tid"), event.get("ts"))] = event

    for flow_id, start_flow in flow_start_dict.items():
        end_flow = flow_end_dict.get(flow_id)
        if end_flow is None:
            continue
        kernel_event = event_dict.get("{}-{}-{}".format(end_flow.get("pid"), end_flow.get("tid"), end_flow.get("ts")))
        if kernel_event is None:
            continue
        flow_kernel_dict.setdefault(start_flow.get("ts"), []).append(kernel_event)
    return event_list, flow_kernel_dict


def get_top_layer_apis(file_path: str, trace_type: int, max_kernel_num: int) -> any:
    event_list, flow_kernel_dict = read_json_file(file_path, trace_type)
    root_node = TreeBuilder.build_tree(event_list, flow_kernel_dict)
    if max_kernel_num is not None:
        TreeBuilder.mark_kernel_num(root_node, flow_kernel_dict)
    level1_child_nodes = root_node.child_nodes
    if not level1_child_nodes:
        raise RuntimeError(f"Can't find any torch op in the file: {file_path}")
    result_data = []
    for level1_node in level1_child_nodes:
        if level1_node.is_step_profiler():
            result_data.extend(level1_node.child_nodes)
        else:
            result_data.append(level1_node)
    return result_data


def compare(base_top_layer_apis: list, comparison_top_layer_apis: list, op_name_map: dict) -> list:
    result_data = []
    comparison_len, base_len = len(comparison_top_layer_apis), len(base_top_layer_apis)
    dp = [[0] * (base_len + 1) for _ in range(comparison_len + 1)]
    for comparison_index in range(1, comparison_len + 1):
        for base_index in range(1, base_len + 1):
            base_name = base_top_layer_apis[base_index - 1].name
            comparison_name = comparison_top_layer_apis[comparison_index - 1].name
            if op_name_map.get(comparison_name, comparison_name) == op_name_map.get(base_name, base_name):
                dp[comparison_index][base_index] = dp[comparison_index - 1][base_index - 1] + 1
            else:
                dp[comparison_index][base_index] = max(dp[comparison_index][base_index - 1],
                                                       dp[comparison_index - 1][base_index])
    matched_op = []
    comparison_index, base_index = comparison_len, base_len
    while comparison_index > 0 and base_index > 0:
        base_name = base_top_layer_apis[base_index - 1].name
        comparison_name = comparison_top_layer_apis[comparison_index - 1].name
        if op_name_map.get(comparison_name, comparison_name) == op_name_map.get(base_name, base_name):
            matched_op.append([comparison_index - 1, base_index - 1])
            comparison_index -= 1
            base_index -= 1
            continue
        if dp[comparison_index][base_index - 1] > dp[comparison_index - 1][base_index]:
            base_index -= 1
        else:
            comparison_index -= 1
    if not matched_op:
        matched_base_index_list = []
    else:
        matched_op.reverse()
        matched_op = np.array(matched_op)
        matched_base_index_list = list(matched_op[:, 1])
    curr_comparison_index = 0
    for base_index, base_api_node in enumerate(base_top_layer_apis):
        if base_index not in matched_base_index_list:
            result_data.append([base_api_node, None])
            continue
        matched_comparison_index = matched_op[matched_base_index_list.index(base_index), 0]
        for comparison_index in range(curr_comparison_index, matched_comparison_index):
            result_data.append([None, comparison_top_layer_apis[comparison_index]])
        result_data.append([base_api_node, comparison_top_layer_apis[matched_comparison_index]])
        curr_comparison_index = matched_comparison_index + 1
    if curr_comparison_index < len(comparison_top_layer_apis):
        for comparison_index in range(curr_comparison_index, len(comparison_top_layer_apis)):
            result_data.append([None, comparison_top_layer_apis[comparison_index]])
    return result_data


def create_data(base_api_node: TorchOpNode, comparison_api_node: TorchOpNode) -> list:
    result_data = []
    base_kernel_list = TreeBuilder.get_total_kernels(base_api_node) if base_api_node else []
    comparison_kernel_list = TreeBuilder.get_total_kernels(comparison_api_node) if comparison_api_node else []
    if not base_kernel_list or not comparison_kernel_list:
        diff = NA
    else:
        base_total_dur = sum([kernel.get("dur", 0) for kernel in base_kernel_list])
        comparison_total_dur = sum([kernel.get("dur", 0) for kernel in comparison_kernel_list])
        diff = (comparison_total_dur - base_total_dur) / base_total_dur
    op_name = base_api_node.name if base_api_node else comparison_api_node.name
    base_kernel_num, comparison_kernel_num = len(base_kernel_list), len(comparison_kernel_list)
    base_data = [NA] * len(GPU_HEADER) if BASE_TRACE_TYPE == GPU else [NA] * len(NPU_HEADER)
    if base_api_node:
        base_data[0] = base_api_node.name
        base_data[1] = base_api_node.input_shape
        base_data[2] = base_api_node.input_type
    comparison_data = [NA] * len(GPU_HEADER) if COMPARISON_TRACE_TYPE == GPU else [NA] * len(NPU_HEADER)
    if comparison_api_node:
        comparison_data[0] = comparison_api_node.name
        comparison_data[1] = comparison_api_node.input_shape
        comparison_data[2] = comparison_api_node.input_type
    if base_kernel_num == 0 and comparison_kernel_num == 0:
        data = base_data + comparison_data + [diff, op_name]
        result_data.append(data)
        return result_data
    for index in range(max(base_kernel_num, comparison_kernel_num)):
        base_row_data, comparison_row_data = copy.deepcopy(base_data), copy.deepcopy(comparison_data)
        if index < base_kernel_num:
            base_kernel = base_kernel_list[index]
            if BASE_TRACE_TYPE == GPU:
                base_row_data[3] = base_kernel.get("name")
                base_row_data[4] = base_kernel.get("dur")
            else:
                base_row_data[3] = base_kernel.get("name")
                base_row_data[4] = base_kernel.get("args", {}).get("Task Id")
                base_row_data[5] = base_kernel.get("args", {}).get("Task Type")
                base_row_data[6] = base_kernel.get("dur")
        if index < comparison_kernel_num:
            comparison_kernel = comparison_kernel_list[index]
            if COMPARISON_TRACE_TYPE == GPU:
                comparison_row_data[3] = comparison_kernel.get("name")
                comparison_row_data[4] = comparison_kernel.get("dur")
            else:
                comparison_row_data[3] = comparison_kernel.get("name")
                comparison_row_data[4] = comparison_kernel.get("args", {}).get("Task Id")
                comparison_row_data[5] = comparison_kernel.get("args", {}).get("Task Type")
                comparison_row_data[6] = comparison_kernel.get("dur")
        data = base_row_data + comparison_row_data + [diff, op_name]
        result_data.append(data)
    return result_data


def drill_down(compare_result_data: list, max_kernel_num: int, op_name_map: dict) -> list:
    result_data = []
    for data in compare_result_data:
        base_api = data[0] if data[0] else TorchOpNode()
        comparison_api = data[1] if data[1] else TorchOpNode()
        if max(base_api.kernel_num, comparison_api.kernel_num) <= max_kernel_num:
            result_data.append(data)
            continue
        result_data.extend(compare(base_api.child_nodes, comparison_api.child_nodes, op_name_map))
    return result_data


def have_to_drill_down(compare_result_data: list, max_kernel_num: int) -> bool:
    for data in compare_result_data:
        base_api = data[0] if data[0] else TorchOpNode()
        comparison_api = data[1] if data[1] else TorchOpNode()
        if max(base_api.kernel_num, comparison_api.kernel_num) > max_kernel_num:
            return True
    return False


def main():
    global BASE_TRACE, COMPARISON_TRACE
    parser = argparse.ArgumentParser(description="Compare trace of GPU and NPU")
    parser.add_argument("base_trace_path", help="base trace file path")
    parser.add_argument("comparison_trace_path", help="comparison trace file path")
    parser.add_argument("--output_path", help="性能数据比对结果的存放路径")
    parser.add_argument("--max_kernel_num", type=int, help="每个torch op的kernel数量限制")
    parser.add_argument("--op_name_map", type=ast.literal_eval, default={},
                        help="配置GPU OP与NPU OP等价的名称映射关系，以字典的形式传入")
    args = parser.parse_args()
    if args.max_kernel_num is not None and args.max_kernel_num <= LIMIT_KERNEL:
        raise RuntimeError(f"Invalid param, --max_kernel_num has to be greater than {LIMIT_KERNEL}")
    if not isinstance(args.op_name_map, dict):
        raise RuntimeError("Invalid param, --op_name_map must be dict, for example: --op_name_map={'name1':'name2'}")
    base_top_layer_apis = get_top_layer_apis(args.base_trace_path, BASE_TYPE, args.max_kernel_num)
    if BASE_TRACE_TYPE == GPU:
        BASE_TRACE += ' [GPU] : ' + os.path.basename(args.base_trace_path)
    else:
        BASE_TRACE += ' [NPU] : ' + os.path.basename(args.base_trace_path)
    comparison_top_layer_apis = get_top_layer_apis(args.comparison_trace_path, COMPARISON_TYPE, args.max_kernel_num)
    if COMPARISON_TRACE_TYPE == GPU:
        COMPARISON_TRACE += ' [GPU] : ' + os.path.basename(args.comparison_trace_path)
    else:
        COMPARISON_TRACE += ' [NPU] : ' + os.path.basename(args.comparison_trace_path)
    compare_result_data = compare(base_top_layer_apis, comparison_top_layer_apis, args.op_name_map)

    if args.max_kernel_num is not None:
        while have_to_drill_down(compare_result_data, args.max_kernel_num):
            compare_result_data = drill_down(compare_result_data, args.max_kernel_num, args.op_name_map)

    dir_path = args.output_path if args.output_path else "./"
    file_name = "torch_op_compare_{}.xlsx".format(time.strftime("%Y%m%d%H%M%S", time.localtime(time.time())))
    result_file_path = os.path.join(dir_path, file_name)

    wb = Workbook()
    ws = wb.create_sheet("CompareResult", 0)
    ws.sheet_properties.tabColor = '00CED1'
    # write headers
    base_trace_headers = GPU_HEADER if BASE_TRACE_TYPE == GPU else NPU_HEADER
    comparison_trace_headers = GPU_HEADER if COMPARISON_TRACE_TYPE == GPU else NPU_HEADER
    headers = base_trace_headers + comparison_trace_headers + [DIFF, OP_NAME_FILTER, DIFF_FILTER]
    base_trace_start_column = 0
    comparison_trace_start_column = len(base_trace_headers)
    diff_start_column = len(base_trace_headers) + len(comparison_trace_headers)

    for col_index in range(len(headers)):
        ws.cell(row=1, column=col_index + 1).border = BORDER
        ws.cell(row=1, column=col_index + 1).font = Font(name='Arial', bold=True)
        ws.cell(row=2, column=col_index + 1).border = BORDER
        ws.cell(row=2, column=col_index + 1).font = Font(name='Arial', bold=True)
        header_name = headers[col_index]
        if col_index < comparison_trace_start_column:
            ws.cell(row=1, column=col_index + 1).value = BASE_TRACE
            ws.cell(row=1, column=col_index + 1).fill = FILL_DICT.get(BASE_TYPE)
            ws.cell(row=2, column=col_index + 1).fill = FILL_DICT.get(BASE_TYPE)
        elif col_index < diff_start_column:
            ws.cell(row=1, column=col_index + 1).value = COMPARISON_TRACE
            ws.cell(row=1, column=col_index + 1).fill = FILL_DICT.get(COMPARISON_TYPE)
            ws.cell(row=2, column=col_index + 1).fill = FILL_DICT.get(COMPARISON_TYPE)
        else:
            ws.cell(row=1, column=col_index + 1).value = header_name
            ws.cell(row=1, column=col_index + 1).fill = FILL_DICT.get(header_name)
        ws.cell(row=2, column=col_index + 1).value = header_name
        dim = ws.cell(row=2, column=col_index + 1).coordinate
        ws.column_dimensions[dim[0]].width = COLUMN_WIDTH.get(header_name)
    ws.merge_cells(start_row=1, start_column=base_trace_start_column + 1,
                   end_row=1, end_column=comparison_trace_start_column)
    ws.merge_cells(start_row=1, start_column=comparison_trace_start_column + 1,
                   end_row=1, end_column=diff_start_column)
    ws.merge_cells(start_row=1, start_column=headers.index(DIFF) + 1,
                   end_row=2, end_column=headers.index(DIFF) + 1)
    ws.merge_cells(start_row=1, start_column=headers.index(OP_NAME_FILTER) + 1,
                   end_row=2, end_column=headers.index(OP_NAME_FILTER) + 1)
    ws.merge_cells(start_row=1, start_column=headers.index(DIFF_FILTER) + 1,
                   end_row=2, end_column=headers.index(DIFF_FILTER) + 1)

    # write lines
    start_row_index = 3
    for data in compare_result_data:
        rows = create_data(data[0], data[1])
        row_number = 0
        for row in rows:
            row_index = start_row_index + row_number
            ws.cell(row=row_index, column=len(row) + 1).border = BORDER
            for index, value in enumerate(row):
                if index == headers.index(DIFF):
                    ws.cell(row=row_index, column=index + 1).number_format = '0.00%'
                    if value != NA and value < 0:
                        ws.cell(row=row_index, column=index + 1).fill = PatternFill("solid", fgColor='0000FF00')
                        ws.cell(row=row_index, column=index + 3).fill = PatternFill("solid", fgColor='0000FF00')
                    if value != NA and value >= 0:
                        ws.cell(row=row_index, column=index + 1).fill = PatternFill("solid", fgColor='00FF0000')
                        ws.cell(row=row_index, column=index + 3).fill = PatternFill("solid", fgColor='00FF0000')
                if index in [key for key, value in enumerate(headers) if value == OP_NAME]:
                    ws.cell(row=row_index, column=index + 1).font = Font(name='Arial', bold=True)
                else:
                    ws.cell(row=row_index, column=index + 1).font = Font(name='Arial')
                ws.cell(row=row_index, column=index + 1).value = value
                ws.cell(row=row_index, column=index + 1).border = BORDER
            row_number += 1
        if row_number > 1:
            # 合并单元格
            merged_index = set(
                [key for key, value in enumerate(headers) if value in (OP_NAME, INPUT_SHAPE, INPUT_TYPE, DIFF)])
            for col_index in merged_index:
                ws.merge_cells(start_row=start_row_index, start_column=col_index + 1,
                               end_row=start_row_index + row_number - 1, end_column=col_index + 1)
        start_row_index = start_row_index + row_number

    wb.save(result_file_path)
    wb.close()


if __name__ == "__main__":
    main()
